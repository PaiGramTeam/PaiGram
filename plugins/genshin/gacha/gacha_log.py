import json
from datetime import datetime
from enum import Enum
from io import BytesIO
from os import sep

import genshin
from genshin.models import BannerType
from modules.apihelper.gacha_log import GachaLog as GachaLogService
from openpyxl import load_workbook
from telegram import Update, User, Message, Document, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ChatAction
from telegram.ext import CallbackContext, CommandHandler, MessageHandler, filters, ConversationHandler

from core.base.assets import AssetsService
from core.baseplugin import BasePlugin
from core.cookies import CookiesService
from core.cookies.error import CookiesNotFoundError
from core.plugin import Plugin, handler, conversation
from core.template import TemplateService
from core.user import UserService
from core.user.error import UserNotFoundError
from modules.apihelper.hyperion import SignIn
from utils.bot import get_all_args
from utils.decorators.admins import bot_admins_rights_check
from utils.decorators.error import error_callable
from utils.decorators.restricts import restricts
from utils.helpers import get_genshin_client
from utils.log import logger
from utils.models.base import RegionEnum

INPUT_URL, INPUT_FILE, CONFIRM_DELETE = range(10100, 10103)


class GachaLog(Plugin.Conversation, BasePlugin.Conversation):
    """抽卡记录导入/导出/分析"""

    def __init__(
        self,
        template_service: TemplateService = None,
        user_service: UserService = None,
        assets: AssetsService = None,
        cookie_service: CookiesService = None,
    ):
        self.template_service = template_service
        self.user_service = user_service
        self.assets_service = assets
        self.cookie_service = cookie_service

    @staticmethod
    def from_url_get_authkey(url: str) -> str:
        """从 UEL 解析 authkey
        :param url: URL
        :return: authkey
        """
        try:
            return url.split("authkey=")[1].split("&")[0]
        except IndexError:
            return url

    @staticmethod
    async def _refresh_user_data(user: User, data: dict = None, authkey: str = None, verify_uid: bool = True) -> str:
        """刷新用户数据
        :param user: 用户
        :param data: 数据
        :param authkey: 认证密钥
        :return: 返回信息
        """
        try:
            logger.debug("尝试获取已绑定的原神账号")
            client = await get_genshin_client(user.id, need_cookie=False)
            if authkey:
                return await GachaLogService.get_gacha_log_data(user.id, client, authkey)
            if data:
                return await GachaLogService.import_gacha_log_data(user.id, client, data, verify_uid)
        except UserNotFoundError:
            logger.info(f"未查询到用户({user.full_name} {user.id}) 所绑定的账号信息")
            return "派蒙没有找到您所绑定的账号信息，请先私聊派蒙绑定账号"

    @staticmethod
    def convert_paimonmoe_to_uigf(data: BytesIO) -> dict:
        """转换 paimone.moe 或 非小酋 导出 xlsx 数据为 UIGF 格式
        :param data: paimon.moe 导出的 xlsx 数据
        :return: UIGF 格式数据
        """
        PAIMONMOE_VERSION = 3
        PM2UIGF_VERSION = 1
        PM2UIGF_NAME = "paimon_moe_to_uigf"
        UIGF_VERSION = "v2.2"

        with open(f"resources{sep}json{sep}zh.json", "r") as load_f:
            zh_dict = json.load(load_f)

        class XlsxType(Enum):
            PAIMONMOE = 1
            FXQ = 2

        class ItemType(Enum):
            CHARACTER = "角色"
            WEAPON = "武器"

        class UIGFGachaType(Enum):
            BEGINNER = 100
            STANDARD = 200
            CHARACTER = 301
            WEAPON = 302

        class Qiyr:
            def __init__(
                self, uigf_gacha_type: UIGFGachaType, item_type: ItemType, name: str, time: datetime, p: int, _id: int
            ) -> None:
                self.uigf_gacha_type = uigf_gacha_type
                self.item_type = item_type
                self.name = name
                self.time = time
                self.rank_type = p
                self.id = _id

            def qy2_json(self):
                return {
                    "gacha_type": self.uigf_gacha_type.value,  # 注意！
                    "item_id": "",
                    "count": -1,
                    "time": self.time.strftime("%Y-%m-%d %H:%M:%S"),
                    "name": self.name,
                    "item_type": self.item_type.value,
                    "rank_type": self.rank_type,
                    "id": self.id,
                    "uigf_gacha_type": self.uigf_gacha_type.value,
                }

        def from_paimon_moe(uigf_gacha_type: UIGFGachaType, item_type: str, name: str, time: str, p: int) -> Qiyr:
            item_type = ItemType.CHARACTER if item_type == "Character" else ItemType.WEAPON
            name = zh_dict[name]

            time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
            return Qiyr(uigf_gacha_type, item_type, name, time, p, 0)

        def from_fxq(uigf_gacha_type: UIGFGachaType, item_type: str, name: str, time: str, p: int, _id: int) -> Qiyr:
            item_type = ItemType.CHARACTER if item_type == "角色" else ItemType.WEAPON
            time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
            return Qiyr(uigf_gacha_type, item_type, name, time, p, _id)

        class uigf:
            qiyes: list[Qiyr]
            uid: int
            export_time: datetime
            export_app: str = PM2UIGF_NAME
            export_app_version: str = PM2UIGF_VERSION
            uigf_version = UIGF_VERSION
            lang = "zh-cn"

            def __init__(self, qiyes: list[Qiyr], uid: int, export_time: datetime) -> None:
                self.uid = uid
                self.qiyes = qiyes
                self.qiyes.sort(key=lambda x: x.time)
                if self.qiyes[0].id == 0:  # 如果是从paimon.moe导入的，那么就给id赋值
                    for index, _ in enumerate(self.qiyes):
                        self.qiyes[index].id = index + 1
                    self.export_time = export_time
                self.export_time = export_time

            def export_json(self) -> dict:
                json_d = {
                    "info": {
                        "uid": self.uid,
                        "lang": self.lang,
                        "export_time": self.export_time.strftime("%Y-%m-%d %H:%M:%S"),
                        "export_timestamp": self.export_time.timestamp(),
                        "export_app": self.export_app,
                        "export_app_version": self.export_app_version,
                        "uigf_version": self.uigf_version,
                    },
                    "list": [],
                }
                for qiye in self.qiyes:
                    json_d["list"].append(qiye.qy2_json())
                return json_d

        wb = load_workbook(data)

        xlsx_type = XlsxType.PAIMONMOE if len(wb.worksheets) == 6 else XlsxType.FXQ  # 判断是paimon.moe还是非小酋导出的

        paimonmoe_sheets = {
            UIGFGachaType.BEGINNER: "Beginners' Wish",
            UIGFGachaType.STANDARD: "Standard",
            UIGFGachaType.CHARACTER: "Character Event",
            UIGFGachaType.WEAPON: "Weapon Event",
        }
        fxq_sheets = {
            UIGFGachaType.BEGINNER: "新手祈愿",
            UIGFGachaType.STANDARD: "常驻祈愿",
            UIGFGachaType.CHARACTER: "角色活动祈愿",
            UIGFGachaType.WEAPON: "武器活动祈愿",
        }
        qiyes = []
        if xlsx_type == XlsxType.PAIMONMOE:
            ws = wb["Information"]
            if ws["B2"].value != PAIMONMOE_VERSION:
                raise Exception("PaimonMoe version not supported")
            export_time = datetime.strptime(ws["B3"].value, "%Y-%m-%d %H:%M:%S")
            for gacha_type in paimonmoe_sheets:
                ws = wb[paimonmoe_sheets[gacha_type]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    qiyes.append(from_paimon_moe(gacha_type, row[0], row[1], row[2], row[3]))
        else:
            export_time = datetime.now()
            for gacha_type in fxq_sheets:
                ws = wb[fxq_sheets[gacha_type]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    qiyes.append(from_fxq(gacha_type, row[2], row[1], row[0], row[3], row[6]))

        u = uigf(qiyes, 0, export_time)
        return u.export_json()

    async def import_from_file(self, user: User, message: Message, document: Document = None) -> None:
        if not document:
            document = message.document
        # TODO: 使用 mimetype 判断文件类型
        if document.file_name.endswith(".xlsx"):
            file_type = "xlsx"
        elif document.file_name.endswith(".json"):
            file_type = "json"
        else:
            await message.reply_text("文件格式错误，请发送符合 UIGF 标准的 json 格式的抽卡记录文件或者 paimon.moe、非小酋导出的 xlsx 格式的抽卡记录文件")
        if document.file_size > 2 * 1024 * 1024:
            await message.reply_text("文件过大，请发送小于 2 MB 的文件")
        try:
            data = BytesIO()
            await (await document.get_file()).download(out=data)
            if file_type == "json":
                # bytesio to json
                data = data.getvalue().decode("utf-8")
                data = json.loads(data)
            else:
                data = self.convert_paimonmoe_to_uigf(data)
        except UnicodeDecodeError:
            await message.reply_text("文件解析失败，请检查文件编码是否正确或符合 UIGF 标准")
            return
        except Exception as exc:
            logger.error(f"文件解析失败：{repr(exc)}")
            await message.reply_text("文件解析失败，请检查文件是否符合 UIGF 标准")
            return
        await message.reply_chat_action(ChatAction.TYPING)
        reply = await message.reply_text("文件解析成功，正在导入数据")
        await message.reply_chat_action(ChatAction.TYPING)
        try:
            text = await self._refresh_user_data(user, data=data, verify_uid=file_type == "json")
        except Exception as exc:  # pylint: disable=W0703
            logger.error(f"文件解析失败：{repr(exc)}")
            text = "文件解析失败，请检查文件是否符合 UIGF 标准"
        await reply.edit_text(text)

    @conversation.entry_point
    @handler(CommandHandler, command="gacha_log_import", filters=filters.ChatType.PRIVATE, block=False)
    @handler(MessageHandler, filters=filters.Regex("^导入抽卡记录(.*)") & filters.ChatType.PRIVATE, block=False)
    @restricts()
    @error_callable
    async def command_start(self, update: Update, context: CallbackContext) -> int:
        message = update.effective_message
        user = update.effective_user
        args = get_all_args(context)
        logger.info(f"用户 {user.full_name}[{user.id}] 导入抽卡记录命令请求")
        authkey = self.from_url_get_authkey(args[0] if args else "")
        if not args:
            if message.document:
                await self.import_from_file(user, message)
                return ConversationHandler.END
            elif message.reply_to_message and message.reply_to_message.document:
                await self.import_from_file(user, message, document=message.reply_to_message.document)
                return ConversationHandler.END
            try:
                user_info = await self.user_service.get_user_by_id(user.id)
            except UserNotFoundError:
                user_info = None
            if user_info and user_info.region == RegionEnum.HYPERION:
                try:
                    cookies = await self.cookie_service.get_cookies(user_info.user_id, user_info.region)
                except CookiesNotFoundError:
                    cookies = None
                if cookies and cookies.cookies and "stoken" in cookies.cookies:
                    if stuid := next(
                        (value for key, value in cookies.cookies.items() if key in ["ltuid", "login_uid"]), None
                    ):
                        cookies.cookies["stuid"] = stuid
                        client = genshin.Client(
                            cookies=cookies.cookies,
                            game=genshin.types.Game.GENSHIN,
                            region=genshin.Region.CHINESE,
                            lang="zh-cn",
                            uid=user_info.yuanshen_uid,
                        )
                        authkey = await SignIn.get_authkey_by_stoken(client)
        if not authkey:
            await message.reply_text(
                "<b>开始导入祈愿历史记录：请通过 https://paimon.moe/wish/import 获取抽卡记录链接后发送给我"
                "（非 paimon.moe 导出的文件数据）</b>\n\n"
                "> 你还可以向派蒙发送从其他工具导出的 UIGF JSON 标准的记录文件\n"
                "> 在绑定 Cookie 时添加 stoken 可能有特殊效果哦（仅限国服）\n"
                "<b>注意：导入的数据将会与旧数据进行合并。</b>",
                parse_mode="html",
            )
            return INPUT_URL
        text = "小派蒙正在从米哈游服务器获取数据，请稍后"
        if not args:
            text += "\n\n> 由于你绑定的 Cookie 中存在 stoken ，本次通过 stoken 自动刷新数据"
        reply = await message.reply_text(text)
        await message.reply_chat_action(ChatAction.TYPING)
        data = await self._refresh_user_data(user, authkey=authkey)
        await reply.edit_text(data)
        return ConversationHandler.END

    @conversation.state(state=INPUT_URL)
    @handler.message(filters=~filters.COMMAND, block=False)
    @restricts()
    @error_callable
    async def import_data_from_message(self, update: Update, _: CallbackContext) -> int:
        message = update.effective_message
        user = update.effective_user
        if message.document:
            await self.import_from_file(user, message)
            return ConversationHandler.END
        authkey = self.from_url_get_authkey(message.text)
        reply = await message.reply_text("小派蒙正在从米哈游服务器获取数据，请稍后")
        await message.reply_chat_action(ChatAction.TYPING)
        text = await self._refresh_user_data(user, authkey=authkey)
        await reply.edit_text(text)
        return ConversationHandler.END

    @conversation.entry_point
    @handler(CommandHandler, command="gacha_log_delete", filters=filters.ChatType.PRIVATE, block=False)
    @handler(MessageHandler, filters=filters.Regex("^删除抽卡记录(.*)") & filters.ChatType.PRIVATE, block=False)
    @restricts()
    @error_callable
    async def command_start_delete(self, update: Update, context: CallbackContext) -> int:
        message = update.effective_message
        user = update.effective_user
        logger.info(f"用户 {user.full_name}[{user.id}] 删除抽卡记录命令请求")
        try:
            client = await get_genshin_client(user.id, need_cookie=False)
            context.chat_data["uid"] = client.uid
        except UserNotFoundError:
            await message.reply_text("你还没有导入抽卡记录哦~")
            return ConversationHandler.END
        _, status = await GachaLogService.load_history_info(str(user.id), str(client.uid), only_status=True)
        if not status:
            await message.reply_text("你还没有导入抽卡记录哦~")
            return ConversationHandler.END
        await message.reply_text("你确定要删除抽卡记录吗？（此项操作无法恢复），如果确定请发送 ”确定“，发送其他内容取消")
        return CONFIRM_DELETE

    @conversation.state(state=CONFIRM_DELETE)
    @handler.message(filters=filters.TEXT & ~filters.COMMAND, block=False)
    @restricts()
    @error_callable
    async def command_confirm_delete(self, update: Update, context: CallbackContext) -> int:
        message = update.effective_message
        user = update.effective_user
        if message.text == "确定":
            status = await GachaLogService.remove_history_info(str(user.id), str(context.chat_data["uid"]))
            await message.reply_text("抽卡记录已删除" if status else "抽卡记录删除失败")
            return ConversationHandler.END
        await message.reply_text("已取消")
        return ConversationHandler.END

    @handler(CommandHandler, command="gacha_log_force_delete", block=False)
    @bot_admins_rights_check
    async def command_gacha_log_force_delete(self, update: Update, context: CallbackContext):
        message = update.effective_message
        args = get_all_args(context)
        if not args:
            await message.reply_text("请指定用户ID")
            return
        try:
            cid = int(args[0])
            if cid < 0:
                raise ValueError("Invalid cid")
            client = await get_genshin_client(cid, need_cookie=False)
            _, status = await GachaLogService.load_history_info(str(cid), str(client.uid), only_status=True)
            if not status:
                await message.reply_text("该用户还没有导入抽卡记录")
                return
            status = await GachaLogService.remove_history_info(str(cid), str(client.uid))
            await message.reply_text("抽卡记录已强制删除" if status else "抽卡记录删除失败")
        except UserNotFoundError:
            await message.reply_text("该用户暂未绑定账号")
        except (ValueError, IndexError):
            await message.reply_text("用户ID 不合法")

    @handler(CommandHandler, command="gacha_log_export", filters=filters.ChatType.PRIVATE, block=False)
    @handler(MessageHandler, filters=filters.Regex("^导出抽卡记录(.*)") & filters.ChatType.PRIVATE, block=False)
    @restricts()
    @error_callable
    async def command_start_export(self, update: Update, context: CallbackContext) -> None:
        message = update.effective_message
        user = update.effective_user
        logger.info(f"用户 {user.full_name}[{user.id}] 导出抽卡记录命令请求")
        try:
            client = await get_genshin_client(user.id, need_cookie=False)
            await message.reply_chat_action(ChatAction.TYPING)
            state, text, path = await GachaLogService.gacha_log_to_uigf(str(user.id), str(client.uid))
            if state:
                await message.reply_chat_action(ChatAction.UPLOAD_DOCUMENT)
                await message.reply_document(document=open(path, "rb+"), caption="抽卡记录导出文件")
            else:
                await message.reply_text(text)
        except UserNotFoundError:
            logger.info(f"未查询到用户({user.full_name} {user.id}) 所绑定的账号信息")
            if filters.ChatType.GROUPS.filter(message):
                buttons = [[InlineKeyboardButton("点我私聊", url=f"https://t.me/{context.bot.username}?start=set_uid")]]
                await message.reply_text("未查询到您所绑定的账号信息，请先私聊派蒙绑定账号", reply_markup=InlineKeyboardMarkup(buttons))
            else:
                await message.reply_text("未查询到您所绑定的账号信息，请先绑定账号")

    @handler(CommandHandler, command="gacha_log", block=False)
    @handler(MessageHandler, filters=filters.Regex("^抽卡记录(.*)"), block=False)
    @restricts()
    @error_callable
    async def command_start_analysis(self, update: Update, context: CallbackContext) -> None:
        message = update.effective_message
        user = update.effective_user
        pool_type = BannerType.CHARACTER1
        if args := get_all_args(context):
            if "武器" in args:
                pool_type = BannerType.WEAPON
            elif "常驻" in args:
                pool_type = BannerType.STANDARD
        logger.info(f"用户 {user.full_name}[{user.id}] 抽卡记录命令请求 || 参数 {pool_type.name}")
        try:
            client = await get_genshin_client(user.id, need_cookie=False)
            await message.reply_chat_action(ChatAction.TYPING)
            data = await GachaLogService.get_analysis(user.id, client, pool_type, self.assets_service)
            if isinstance(data, str):
                reply_message = await message.reply_text(data)
                if filters.ChatType.GROUPS.filter(message):
                    self._add_delete_message_job(context, reply_message.chat_id, reply_message.message_id, 300)
                    self._add_delete_message_job(context, message.chat_id, message.message_id, 300)
            else:
                await message.reply_chat_action(ChatAction.UPLOAD_PHOTO)
                png_data = await self.template_service.render(
                    "genshin/gacha_log/gacha_log.html", data, full_page=True, query_selector=".body_box"
                )
                await message.reply_photo(png_data)
        except UserNotFoundError:
            logger.info(f"未查询到用户({user.full_name} {user.id}) 所绑定的账号信息")
            if filters.ChatType.GROUPS.filter(message):
                buttons = [[InlineKeyboardButton("点我私聊", url=f"https://t.me/{context.bot.username}?start=set_cookie")]]
                reply_message = await message.reply_text(
                    "未查询到您所绑定的账号信息，请先私聊派蒙绑定账号", reply_markup=InlineKeyboardMarkup(buttons)
                )
                self._add_delete_message_job(context, reply_message.chat_id, reply_message.message_id, 30)

                self._add_delete_message_job(context, message.chat_id, message.message_id, 30)
            else:
                await message.reply_text("未查询到您所绑定的账号信息，请先绑定账号")

    @handler(CommandHandler, command="gacha_count", block=True)
    @handler(MessageHandler, filters=filters.Regex("^抽卡统计(.*)"), block=True)
    @restricts()
    @error_callable
    async def command_start_count(self, update: Update, context: CallbackContext) -> None:
        message = update.effective_message
        user = update.effective_user
        pool_type = BannerType.CHARACTER1
        if args := get_all_args(context):
            if "武器" in args:
                pool_type = BannerType.WEAPON
            elif "常驻" in args:
                pool_type = BannerType.STANDARD
        logger.info(f"用户 {user.full_name}[{user.id}] 抽卡统计命令请求 || 参数 {pool_type.name}")
        try:
            client = await get_genshin_client(user.id, need_cookie=False)
            group = filters.ChatType.GROUPS.filter(message)
            await message.reply_chat_action(ChatAction.TYPING)
            data = await GachaLogService.get_pool_analysis(user.id, client, pool_type, self.assets_service, group)
            if isinstance(data, str):
                reply_message = await message.reply_text(data)
                if filters.ChatType.GROUPS.filter(message):
                    self._add_delete_message_job(context, reply_message.chat_id, reply_message.message_id, 300)
                    self._add_delete_message_job(context, message.chat_id, message.message_id, 300)
            else:
                document = False
                if data["hasMore"] and not group:
                    document = True
                    data["hasMore"] = False
                png_data = await self.template_service.render(
                    "genshin/gacha_count/gacha_count.html", data, full_page=True, query_selector=".body_box"
                )
                if document:
                    await message.reply_chat_action(ChatAction.UPLOAD_DOCUMENT)
                    await message.reply_document(png_data, filename="抽卡统计.png")
                else:
                    await message.reply_chat_action(ChatAction.UPLOAD_PHOTO)
                    await message.reply_photo(png_data)
        except (UserNotFoundError, CookiesNotFoundError):
            logger.info(f"未查询到用户({user.full_name} {user.id}) 所绑定的账号信息")
            if filters.ChatType.GROUPS.filter(message):
                buttons = [[InlineKeyboardButton("点我私聊", url=f"https://t.me/{context.bot.username}?start=set_cookie")]]
                reply_message = await message.reply_text(
                    "未查询到您所绑定的账号信息，请先私聊派蒙绑定账号", reply_markup=InlineKeyboardMarkup(buttons)
                )
                self._add_delete_message_job(context, reply_message.chat_id, reply_message.message_id, 30)

                self._add_delete_message_job(context, message.chat_id, message.message_id, 30)
            else:
                await message.reply_text("未查询到您所绑定的账号信息，请先绑定账号")
