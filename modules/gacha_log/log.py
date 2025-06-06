import asyncio
import contextlib
import datetime
import json
from concurrent.futures import ThreadPoolExecutor
from os import PathLike
from pathlib import Path
from typing import Dict, IO, List, Optional, Tuple, Union, TYPE_CHECKING

import aiofiles
from openpyxl import load_workbook
from simnet import GenshinClient, Region
from simnet.errors import AuthkeyTimeout, InvalidAuthkey
from simnet.models.base import add_timezone
from simnet.models.genshin.wish import BannerType
from simnet.utils.player import recognize_genshin_server

from gram_core.services.gacha_log_rank.services import GachaLogRankService
from metadata.pool.pool import get_pool_by_id
from metadata.shortname import roleToId, weaponToId
from modules.gacha_log.const import GACHA_TYPE_LIST, PAIMONMOE_VERSION
from modules.gacha_log.error import (
    GachaLogAccountNotFound,
    GachaLogAuthkeyTimeout,
    GachaLogException,
    GachaLogFileError,
    GachaLogInvalidAuthkey,
    GachaLogMixedProvider,
    GachaLogNotFound,
    PaimonMoeGachaLogFileError,
)
from modules.gacha_log.models import (
    FiveStarItem,
    FourStarItem,
    GachaItem,
    GachaLogInfo,
    ImportType,
    ItemType,
    Pool,
    UIGFGachaType,
    UIGFInfo,
    UIGFItem,
    UIGFModel,
    UIGFListInfo,
)
from modules.gacha_log.online_view import GachaLogOnlineView
from modules.gacha_log.ranks import GachaLogRanks
from modules.gacha_log.uigf import GachaLogUigfConverter
from utils.const import PROJECT_ROOT
from utils.uid import mask_number

if TYPE_CHECKING:
    from core.dependence.assets.impl.genshin import AssetsService


GACHA_LOG_PATH = PROJECT_ROOT.joinpath("data", "apihelper", "gacha_log")
GACHA_LOG_PATH.mkdir(parents=True, exist_ok=True)


class GachaLog(GachaLogOnlineView, GachaLogRanks, GachaLogUigfConverter):
    def __init__(
        self,
        gacha_log_path: Path = GACHA_LOG_PATH,
        gacha_log_rank_service: GachaLogRankService = None,
    ):
        GachaLogOnlineView.__init__(self)
        GachaLogRanks.__init__(self, gacha_log_rank_service)
        self.gacha_log_path = gacha_log_path

    @staticmethod
    async def load_json(path):
        async with aiofiles.open(path, "r", encoding="utf-8") as f:
            return json.loads(await f.read())

    @staticmethod
    async def save_json(path, data):
        async with aiofiles.open(path, "w", encoding="utf-8") as f:
            if isinstance(data, dict):
                return await f.write(json.dumps(data, ensure_ascii=False, indent=4))
            await f.write(data)

    async def load_history_info(
        self, user_id: str, uid: str, only_status: bool = False
    ) -> Tuple[Optional[GachaLogInfo], bool]:
        """读取历史抽卡记录数据
        :param user_id: 用户id
        :param uid: 原神uid
        :param only_status: 是否只读取状态
        :return: 抽卡记录数据
        """
        file_path = self.gacha_log_path / f"{user_id}-{uid}.json"
        if only_status:
            return None, file_path.exists()
        if not file_path.exists():
            return GachaLogInfo(user_id=user_id, uid=uid, update_time=datetime.datetime.now()), False
        try:
            return GachaLogInfo.parse_obj(await self.load_json(file_path)), True
        except json.decoder.JSONDecodeError:
            return GachaLogInfo(user_id=user_id, uid=uid, update_time=datetime.datetime.now()), False

    async def remove_history_info(self, user_id: str, uid: str) -> bool:
        """删除历史抽卡记录数据
        :param user_id: 用户id
        :param uid: 原神uid
        :return: 是否删除成功
        """
        file_path = self.gacha_log_path / f"{user_id}-{uid}.json"
        file_bak_path = self.gacha_log_path / f"{user_id}-{uid}.json.bak"
        file_export_path = self.gacha_log_path / f"{user_id}-{uid}-uigf.json"
        with contextlib.suppress(Exception):
            file_bak_path.unlink(missing_ok=True)
        with contextlib.suppress(Exception):
            file_export_path.unlink(missing_ok=True)
        if file_path.exists():
            try:
                file_path.unlink()
            except PermissionError:
                return False
            return True
        return False

    async def move_history_info(self, user_id: str, uid: str, new_user_id: str) -> bool:
        """移动历史抽卡记录数据
        :param user_id: 用户id
        :param uid: 原神uid
        :param new_user_id: 新用户id
        :return: 是否移动成功
        """
        old_file_path = self.gacha_log_path / f"{user_id}-{uid}.json"
        new_file_path = self.gacha_log_path / f"{new_user_id}-{uid}.json"
        if (not old_file_path.exists()) or new_file_path.exists():
            return False
        try:
            old_file_path.rename(new_file_path)
            return True
        except PermissionError:
            return False

    async def save_gacha_log_info(self, user_id: str, uid: str, info: GachaLogInfo):
        """保存抽卡记录数据
        :param user_id: 用户id
        :param uid: 玩家uid
        :param info: 抽卡记录数据
        """
        save_path = self.gacha_log_path / f"{user_id}-{uid}.json"
        save_path_bak = self.gacha_log_path / f"{user_id}-{uid}.json.bak"
        # 将旧数据备份一次
        with contextlib.suppress(PermissionError):
            if save_path.exists():
                if save_path_bak.exists():
                    save_path_bak.unlink()
                save_path.rename(save_path.parent / f"{save_path.name}.bak")
        # 写入数据
        await self.save_json(save_path, info.json())

    @staticmethod
    async def verify_data(data: List[GachaItem]) -> bool:
        try:
            total = len(data)
            five_star = len([i for i in data if i.rank_type == "5"])
            four_star = len([i for i in data if i.rank_type == "4"])
            if total > 50:
                if total <= five_star * 15:
                    raise GachaLogFileError(
                        "检测到您将要导入的抽卡记录中五星数量过多，可能是由于文件错误导致的，请检查后重新导入。"
                    )
                if four_star < five_star:
                    raise GachaLogFileError(
                        "检测到您将要导入的抽卡记录中五星数量过多，可能是由于文件错误导致的，请检查后重新导入。"
                    )
            return True
        except Exception as exc:  # pylint: disable=W0703
            raise GachaLogFileError from exc

    @staticmethod
    def import_data_backend(all_items: List[GachaItem], gacha_log: GachaLogInfo, temp_id_data: Dict) -> int:
        new_num = 0
        for item_info in all_items:
            pool_name = GACHA_TYPE_LIST[BannerType(int(item_info.gacha_type))]
            if pool_name not in temp_id_data:
                temp_id_data[pool_name] = []
            if pool_name not in gacha_log.item_list:
                gacha_log.item_list[pool_name] = []
            if item_info.id not in temp_id_data[pool_name]:
                gacha_log.item_list[pool_name].append(item_info)
                temp_id_data[pool_name].append(item_info.id)
                new_num += 1
        return new_num

    async def import_gacha_log_data(self, user_id: int, player_id: int, data: dict, verify_uid: bool = True) -> int:
        new_num = 0
        try:
            _data, uid = None, None
            for _i in data.get("hk4e", []):
                uid = _i.get("uid", "0")
                if (not verify_uid) or int(uid) == player_id:
                    _data = _i
                    break
            if not _data or not uid:
                raise GachaLogAccountNotFound
            try:
                import_type = ImportType(data["info"]["export_app"])
            except ValueError:
                import_type = ImportType.UNKNOWN
            # 检查导入数据是否合法
            all_items = [GachaItem(**i) for i in _data["list"]]
            await self.verify_data(all_items)
            gacha_log, status = await self.load_history_info(str(user_id), uid)
            if import_type == ImportType.PAIMONMOE:
                if status and gacha_log.get_import_type != ImportType.PAIMONMOE:
                    raise GachaLogMixedProvider
            elif status and gacha_log.get_import_type == ImportType.PAIMONMOE:
                raise GachaLogMixedProvider
            # 将唯一 id 放入临时数据中，加快查找速度
            temp_id_data = {
                pool_name: [i.id for i in pool_data] for pool_name, pool_data in gacha_log.item_list.items()
            }
            # 使用新线程进行遍历，避免堵塞主线程
            loop = asyncio.get_event_loop()
            # 可以使用with语句来确保线程执行完成后及时被清理
            with ThreadPoolExecutor() as executor:
                new_num = await loop.run_in_executor(
                    executor, self.import_data_backend, all_items, gacha_log, temp_id_data
                )
            for i in gacha_log.item_list.values():
                # 检查导入后的数据是否合法
                await self.verify_data(i)
                i.sort(key=lambda x: (x.time, x.id))
            gacha_log.update_time = add_timezone(datetime.datetime.now())
            gacha_log.import_type = import_type.value
            await self.save_gacha_log_info(str(user_id), uid, gacha_log)
            await self.recount_one_from_uid(user_id, player_id)
            return new_num
        except GachaLogAccountNotFound as e:
            raise GachaLogAccountNotFound("导入失败，文件包含的祈愿记录所属 uid 与你当前绑定的 uid 不同") from e
        except GachaLogMixedProvider as e:
            raise GachaLogMixedProvider from e
        except Exception as exc:
            raise GachaLogException from exc

    @staticmethod
    def get_game_client(player_id: int) -> GenshinClient:
        if recognize_genshin_server(player_id) in ["cn_gf01", "cn_qd01"]:
            return GenshinClient(player_id=player_id, region=Region.CHINESE, lang="zh-cn")
        return GenshinClient(player_id=player_id, region=Region.OVERSEAS, lang="zh-cn")

    async def get_gacha_log_data(self, user_id: int, player_id: int, authkey: str, is_lazy: bool) -> int:
        """使用authkey获取抽卡记录数据，并合并旧数据
        :param user_id: 用户id
        :param player_id: 玩家id
        :param authkey: authkey
        :param is_lazy: 是否快速导入
        :return: 更新结果
        """
        new_num = 0
        gacha_log, _ = await self.load_history_info(str(user_id), str(player_id))
        if gacha_log.get_import_type == ImportType.PAIMONMOE:
            raise GachaLogMixedProvider
        # 将唯一 id 放入临时数据中，加快查找速度
        temp_id_data = {pool_name: [i.id for i in pool_data] for pool_name, pool_data in gacha_log.item_list.items()}
        client = self.get_game_client(player_id)
        try:
            for pool_id, pool_name in GACHA_TYPE_LIST.items():
                if pool_name not in temp_id_data:
                    temp_id_data[pool_name] = []
                if pool_name not in gacha_log.item_list:
                    gacha_log.item_list[pool_name] = []
                min_id = 0
                if is_lazy and gacha_log.item_list[pool_name]:
                    with contextlib.suppress(ValueError):
                        min_id = int(gacha_log.item_list[pool_name][-1].id)

                wish_history = await client.wish_history(pool_id.value, authkey=authkey, min_id=min_id)

                if not is_lazy:
                    min_id = min([i.id for i in wish_history[:20]]) if wish_history else min_id
                    if min_id:
                        gacha_log.item_list[pool_name][:] = filter(
                            lambda i: int(i.id) < min_id, gacha_log.item_list[pool_name]
                        )
                for data in wish_history:
                    item = GachaItem(
                        id=str(data.id),
                        name=data.name,
                        gacha_type=str(data.banner_type.value),
                        item_type=data.type,
                        rank_type=str(data.rarity),
                        time=data.time,
                    )

                    if item.id not in temp_id_data[pool_name] or (not is_lazy and min_id):
                        gacha_log.item_list[pool_name].append(item)
                        temp_id_data[pool_name].append(item.id)
                        new_num += 1

                await asyncio.sleep(1)
        except AuthkeyTimeout as exc:
            raise GachaLogAuthkeyTimeout from exc
        except InvalidAuthkey as exc:
            raise GachaLogInvalidAuthkey from exc
        finally:
            await client.shutdown()
        for i in gacha_log.item_list.values():
            i.sort(key=lambda x: (x.time, x.id))
        gacha_log.update_time = add_timezone(datetime.datetime.now())
        gacha_log.import_type = ImportType.UIGF.value
        await self.save_gacha_log_info(str(user_id), str(player_id), gacha_log)
        await self.recount_one_from_uid(user_id, player_id)
        return new_num

    @staticmethod
    def format_time(time: str) -> datetime.datetime:
        return add_timezone(datetime.datetime.strptime(time, "%Y-%m-%d %H:%M:%S"))

    @staticmethod
    def check_avatar_up(name: str, gacha_time: datetime.datetime) -> bool:
        if name in {"莫娜", "七七", "迪卢克", "琴", "迪希雅"}:
            return False
        if name == "刻晴":
            start_time = GachaLog.format_time("2021-02-17 18:00:00")
            end_time = GachaLog.format_time("2021-03-02 15:59:59")
            if not start_time < gacha_time < end_time:
                return False
        elif name == "提纳里":
            start_time = GachaLog.format_time("2022-08-24 06:00:00")
            end_time = GachaLog.format_time("2022-09-09 17:59:59")
            if not start_time < gacha_time < end_time:
                return False
        elif name == "梦见月瑞希":
            start_time = GachaLog.format_time("2025-02-12 06:00:00")
            end_time = GachaLog.format_time("2025-03-04 17:59:59")
            if not start_time < gacha_time < end_time:
                return False
        return True

    async def get_all_5_star_items(self, data: List[GachaItem], assets: "AssetsService", pool_name: str = "角色祈愿"):
        """
        获取所有5星角色
        :param data: 抽卡记录
        :param assets: 资源服务
        :param pool_name: 池子名称
        :return: 5星角色列表
        """
        count = 0
        result = []
        for item in data:
            count += 1
            if item.rank_type == "5":
                if item.item_type == "角色" and pool_name in {"角色祈愿", "常驻祈愿", "新手祈愿", "集录祈愿"}:
                    data = {
                        "name": item.name,
                        "icon": assets.avatar.icon(roleToId(item.name)).as_uri() if assets else "",
                        "count": count,
                        "type": "角色",
                        "isUp": self.check_avatar_up(item.name, item.time) if pool_name == "角色祈愿" else False,
                        "isBig": (not result[-1].isUp) if result and pool_name == "角色祈愿" else False,
                        "time": item.time,
                    }
                    result.append(FiveStarItem.construct(**data))
                elif item.item_type == "武器" and pool_name in {"武器祈愿", "常驻祈愿", "新手祈愿", "集录祈愿"}:
                    data = {
                        "name": item.name,
                        "icon": assets.weapon.icon(weaponToId(item.name)).as_uri() if assets else "",
                        "count": count,
                        "type": "武器",
                        "isUp": False,
                        "isBig": False,
                        "time": item.time,
                    }
                    result.append(FiveStarItem.construct(**data))
                count = 0
        result.reverse()
        return result, count

    @staticmethod
    async def get_all_4_star_items(data: List[GachaItem], assets: "AssetsService"):
        """
        获取 no_fout_star
        :param data: 抽卡记录
        :param assets: 资源服务
        :return: no_fout_star
        """
        count = 0
        result = []
        for item in data:
            count += 1
            if item.rank_type == "4":
                if item.item_type == "角色":
                    data = {
                        "name": item.name,
                        "icon": assets.avatar.icon(roleToId(item.name)).as_uri() if assets else "",
                        "count": count,
                        "type": "角色",
                        "time": item.time,
                    }
                    result.append(FourStarItem.construct(**data))
                elif item.item_type == "武器":
                    data = {
                        "name": item.name,
                        "icon": assets.weapon.icon(weaponToId(item.name)).as_uri() if assets else "",
                        "count": count,
                        "type": "武器",
                        "time": item.time,
                    }
                    result.append(FourStarItem.construct(**data))
                count = 0
        result.reverse()
        return result, count

    @staticmethod
    def get_301_pool_data(total: int, all_five: List[FiveStarItem], no_five_star: int, no_four_star: int):
        # 总共五星
        five_star = len(all_five)
        five_star_up = len([i for i in all_five if i.isUp])
        five_star_big = len([i for i in all_five if i.isBig])
        # 五星平均
        five_star_avg = round((total - no_five_star) / five_star, 2) if five_star != 0 else 0
        # 小保底不歪
        small_protect = (
            round((five_star_up - five_star_big) / (five_star - five_star_big) * 100.0, 1)
            if five_star - five_star_big != 0
            else "0.0"
        )
        # 五星常驻
        five_star_const = five_star - five_star_up
        # UP 平均
        up_avg = (
            round((total - no_five_star - (all_five[0].count if not all_five[0].isUp else 0)) / five_star_up, 2)
            if five_star_up != 0
            else 0
        )
        # UP 花费原石
        up_cost = sum(i.count * 160 for i in all_five if i.isUp)
        up_cost = f"{round(up_cost / 10000, 2)}w" if up_cost >= 10000 else up_cost
        return [
            [
                {"num": no_five_star, "unit": "抽", "lable": "未出五星"},
                {"num": five_star, "unit": "个", "lable": "五星"},
                {"num": five_star_avg, "unit": "抽", "lable": "五星平均"},
                {"num": small_protect, "unit": "%", "lable": "小保底不歪"},
                {"num": no_four_star, "unit": "抽", "lable": "未出四星"},
                {"num": five_star_const, "unit": "个", "lable": "五星常驻"},
                {"num": up_avg, "unit": "抽", "lable": "UP平均"},
                {"num": up_cost, "unit": "", "lable": "UP花费原石"},
            ],
        ]

    @staticmethod
    def get_200_pool_data(
        total: int, all_five: List[FiveStarItem], all_four: List[FourStarItem], no_five_star: int, no_four_star: int
    ):
        # 总共五星
        five_star = len(all_five)
        # 五星平均
        five_star_avg = round((total - no_five_star) / five_star, 2) if five_star != 0 else 0
        # 五星武器
        five_star_weapon = len([i for i in all_five if i.type == "武器"])
        # 总共四星
        four_star = len(all_four)
        # 四星平均
        four_star_avg = round((total - no_four_star) / four_star, 2) if four_star != 0 else 0
        # 四星最多
        four_star_name_list = [i.name for i in all_four]
        four_star_max = max(four_star_name_list, key=four_star_name_list.count) if four_star_name_list else ""
        four_star_max_count = four_star_name_list.count(four_star_max)
        return [
            [
                {"num": no_five_star, "unit": "抽", "lable": "未出五星"},
                {"num": five_star, "unit": "个", "lable": "五星"},
                {"num": five_star_avg, "unit": "抽", "lable": "五星平均"},
                {"num": five_star_weapon, "unit": "个", "lable": "五星武器"},
                {"num": no_four_star, "unit": "抽", "lable": "未出四星"},
                {"num": four_star, "unit": "个", "lable": "四星"},
                {"num": four_star_avg, "unit": "抽", "lable": "四星平均"},
                {"num": four_star_max_count, "unit": four_star_max, "lable": "四星最多"},
            ],
        ]

    @staticmethod
    def get_302_pool_data(
        total: int, all_five: List[FiveStarItem], all_four: List[FourStarItem], no_five_star: int, no_four_star: int
    ):
        # 总共五星
        five_star = len(all_five)
        # 五星平均
        five_star_avg = round((total - no_five_star) / five_star, 2) if five_star != 0 else 0
        # 四星武器
        four_star_weapon = len([i for i in all_four if i.type == "武器"])
        # 总共四星
        four_star = len(all_four)
        # 四星平均
        four_star_avg = round((total - no_four_star) / four_star, 2) if four_star != 0 else 0
        # 四星最多
        four_star_name_list = [i.name for i in all_four]
        four_star_max = max(four_star_name_list, key=four_star_name_list.count) if four_star_name_list else ""
        four_star_max_count = four_star_name_list.count(four_star_max)
        return [
            [
                {"num": no_five_star, "unit": "抽", "lable": "未出五星"},
                {"num": five_star, "unit": "个", "lable": "五星"},
                {"num": five_star_avg, "unit": "抽", "lable": "五星平均"},
                {"num": four_star_weapon, "unit": "个", "lable": "四星武器"},
                {"num": no_four_star, "unit": "抽", "lable": "未出四星"},
                {"num": four_star, "unit": "个", "lable": "四星"},
                {"num": four_star_avg, "unit": "抽", "lable": "四星平均"},
                {"num": four_star_max_count, "unit": four_star_max, "lable": "四星最多"},
            ],
        ]

    @staticmethod
    def get_500_pool_data(
        total: int, all_five: List[FiveStarItem], all_four: List[FourStarItem], no_five_star: int, no_four_star: int
    ):
        # 总共五星
        five_star = len(all_five)
        # 五星平均
        five_star_avg = round((total - no_five_star) / five_star, 2) if five_star != 0 else 0
        # 四星角色
        four_star_character = len([i for i in all_four if i.type == "角色"])
        # 总共四星
        four_star = len(all_four)
        # 四星平均
        four_star_avg = round((total - no_four_star) / four_star, 2) if four_star != 0 else 0
        # 四星最多
        four_star_name_list = [i.name for i in all_four]
        four_star_max = max(four_star_name_list, key=four_star_name_list.count) if four_star_name_list else ""
        four_star_max_count = four_star_name_list.count(four_star_max)
        return [
            [
                {"num": no_five_star, "unit": "抽", "lable": "未出五星"},
                {"num": five_star, "unit": "个", "lable": "五星"},
                {"num": five_star_avg, "unit": "抽", "lable": "五星平均"},
                {"num": four_star_character, "unit": "个", "lable": "四星角色"},
                {"num": no_four_star, "unit": "抽", "lable": "未出四星"},
                {"num": four_star, "unit": "个", "lable": "四星"},
                {"num": four_star_avg, "unit": "抽", "lable": "四星平均"},
                {"num": four_star_max_count, "unit": four_star_max, "lable": "四星最多"},
            ],
        ]

    @staticmethod
    def count_fortune(pool_name: str, summon_data, weapon: bool = False):
        """
            角色  武器
        欧 50以下 45以下
        吉 50-60 45-55
        中 60-70 55-65
        非 70以上 65以上
        """
        data = [45, 55, 65] if weapon else [50, 60, 70]
        for i in summon_data:
            for j in i:
                if j.get("lable") == "五星平均":
                    num = j.get("num", 0)
                    if num == 0:
                        return pool_name
                    if num <= data[0]:
                        return f"{pool_name} · 欧"
                    if num <= data[1]:
                        return f"{pool_name} · 吉"
                    if num <= data[2]:
                        return f"{pool_name} · 普通"
                    return f"{pool_name} · 非"
        return pool_name

    async def get_analysis(self, user_id: int, player_id: int, pool: BannerType, assets: "AssetsService"):
        """
        获取抽卡记录分析数据
        :param user_id: 用户id
        :param player_id: 玩家id
        :param pool: 池子类型
        :param assets: 资源服务
        :return: 分析数据
        """
        gacha_log, status = await self.load_history_info(str(user_id), str(player_id))
        if not status:
            raise GachaLogNotFound
        return await self.get_analysis_data(gacha_log, pool, assets)

    async def get_analysis_data(self, gacha_log: "GachaLogInfo", pool: BannerType, assets: Optional["AssetsService"]):
        """
        获取抽卡记录分析数据
        :param gacha_log: 抽卡记录
        :param pool: 池子类型
        :param assets: 资源服务
        :return: 分析数据
        """
        player_id = gacha_log.uid
        pool_name = GACHA_TYPE_LIST[pool]
        if pool_name not in gacha_log.item_list:
            raise GachaLogNotFound
        data = gacha_log.item_list[pool_name]
        total = len(data)
        if total == 0:
            raise GachaLogNotFound
        all_five, no_five_star = await self.get_all_5_star_items(data, assets, pool_name)
        all_four, no_four_star = await self.get_all_4_star_items(data, assets)
        summon_data = None
        if pool in [BannerType.CHARACTER1, BannerType.CHARACTER2, BannerType.NOVICE]:
            summon_data = self.get_301_pool_data(total, all_five, no_five_star, no_four_star)
            pool_name = self.count_fortune(pool_name, summon_data)
        elif pool == BannerType.WEAPON:
            summon_data = self.get_302_pool_data(total, all_five, all_four, no_five_star, no_four_star)
            pool_name = self.count_fortune(pool_name, summon_data, True)
        elif pool == BannerType.PERMANENT:
            summon_data = self.get_200_pool_data(total, all_five, all_four, no_five_star, no_four_star)
            pool_name = self.count_fortune(pool_name, summon_data)
        elif pool == BannerType.CHRONICLED:
            summon_data = self.get_500_pool_data(total, all_five, all_four, no_five_star, no_four_star)
            pool_name = self.count_fortune(pool_name, summon_data)
        last_time = data[0].time.strftime("%Y-%m-%d %H:%M")
        first_time = data[-1].time.strftime("%Y-%m-%d %H:%M")
        return {
            "uid": mask_number(player_id),
            "allNum": total,
            "type": pool.value,
            "typeName": pool_name,
            "line": summon_data,
            "firstTime": first_time,
            "lastTime": last_time,
            "fiveLog": all_five,
            "fourLog": all_four[:36],
        }

    async def get_pool_analysis(
        self, user_id: int, player_id: int, pool: BannerType, assets: "AssetsService", group: bool
    ) -> dict:
        """获取抽卡记录分析数据
        :param user_id: 用户id
        :param player_id: 玩家id
        :param pool: 池子类型
        :param assets: 资源服务
        :param group: 是否群组
        :return: 分析数据
        """
        gacha_log, status = await self.load_history_info(str(user_id), str(player_id))
        if not status:
            raise GachaLogNotFound
        pool_name = GACHA_TYPE_LIST[pool]
        if pool_name not in gacha_log.item_list:
            raise GachaLogNotFound
        data = gacha_log.item_list[pool_name]
        total = len(data)
        if total == 0:
            raise GachaLogNotFound
        all_five, _ = await self.get_all_5_star_items(data, assets, pool_name)
        all_four, _ = await self.get_all_4_star_items(data, assets)
        pool_data = []
        up_pool_data = [Pool(**i) for i in get_pool_by_id(pool.value)]
        for up_pool in up_pool_data:
            for item in all_five:
                up_pool.parse(item)
            for item in all_four:
                up_pool.parse(item)
            up_pool.count_item(data)
        for up_pool in up_pool_data:
            pool_data.append(
                {
                    "count": up_pool.count,
                    "list": up_pool.to_list(),
                    "name": up_pool.name,
                    "start": up_pool.start.strftime("%Y-%m-%d"),
                    "end": up_pool.end.strftime("%Y-%m-%d"),
                }
            )
        pool_data = [i for i in pool_data if i["count"] > 0]
        return {
            "uid": player_id,
            "typeName": pool_name,
            "pool": pool_data[:6] if group else pool_data,
            "hasMore": len(pool_data) > 6,
        }

    async def get_all_five_analysis(self, user_id: int, player_id: int, assets: "AssetsService") -> dict:
        """获取五星抽卡记录分析数据
        :param user_id: 用户id
        :param player_id: 玩家id
        :param assets: 资源服务
        :return: 分析数据
        """
        gacha_log, status = await self.load_history_info(str(user_id), str(player_id))
        if not status:
            raise GachaLogNotFound
        pools = []
        for pool_name, items in gacha_log.item_list.items():
            pool = Pool(
                five=[pool_name],
                four=[],
                name=pool_name,
                to=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                **{"from": "2020-09-28 00:00:00"},
            )
            all_five, _ = await self.get_all_5_star_items(items, assets, pool_name)
            for item in all_five:
                pool.parse(item)
            pool.count_item(items)
            pools.append(pool)
        pool_data = [
            {
                "count": up_pool.count,
                "list": up_pool.to_list(),
                "name": up_pool.name,
                "start": up_pool.start.strftime("%Y-%m-%d"),
                "end": up_pool.end.strftime("%Y-%m-%d"),
            }
            for up_pool in pools
        ]
        return {
            "uid": player_id,
            "typeName": "五星列表",
            "pool": pool_data,
            "hasMore": False,
        }

    @staticmethod
    def convert_xlsx_to_uigf(file: Union[str, PathLike, IO[bytes]], zh_dict: Dict) -> Dict:
        """转换 paimone.moe 或 非小酋 导出 xlsx 数据为 UIGF 格式
        :param file: 导出的 xlsx 文件
        :param zh_dict:
        :return: UIGF 格式数据
        """

        def from_paimon_moe(
            uigf_gacha_type: UIGFGachaType, item_type: str, name: str, date_string: str, rank_type: int, _id: int
        ) -> UIGFItem:
            item_type = ItemType.CHARACTER if item_type == "Character" else ItemType.WEAPON
            return UIGFItem(
                id=str(_id),
                name=zh_dict[name],
                gacha_type=uigf_gacha_type,
                item_type=item_type,
                rank_type=str(rank_type),
                time=date_string,
                uigf_gacha_type=uigf_gacha_type,
            )

        def from_uigf(
            uigf_gacha_type: str,
            gacha__type: str,
            item_type: str,
            name: str,
            date_string: str,
            rank_type: str,
            _id: str,
        ) -> UIGFItem:
            return UIGFItem(
                id=_id,
                name=name,
                gacha_type=gacha__type,
                item_type=item_type,
                rank_type=rank_type,
                time=date_string,
                uigf_gacha_type=uigf_gacha_type,
            )

        def from_fxq(
            uigf_gacha_type: UIGFGachaType, item_type: str, name: str, date_string: str, rank_type: int, _id: int
        ) -> UIGFItem:
            item_type = ItemType.CHARACTER if item_type == "角色" else ItemType.WEAPON
            return UIGFItem(
                id=str(_id),
                name=name,
                gacha_type=uigf_gacha_type,
                item_type=item_type,
                rank_type=str(rank_type),
                time=date_string,
                uigf_gacha_type=uigf_gacha_type,
            )

        wb = load_workbook(file)
        wb_len = len(wb.worksheets)

        if wb_len == 6:
            import_type = ImportType.PAIMONMOE
        elif wb_len == 5:
            import_type = ImportType.UIGF
        elif wb_len == 4:
            import_type = ImportType.FXQ
        else:
            raise GachaLogFileError("xlsx 格式错误")

        paimonmoe_sheets = {
            UIGFGachaType.BEGINNER: "Beginners' Wish",
            UIGFGachaType.STANDARD: "Standard",
            UIGFGachaType.CHARACTER: "Character Event",
            UIGFGachaType.WEAPON: "Weapon Event",
        }
        fxq_sheets = {
            UIGFGachaType.BEGINNER: "新手祈愿",
            UIGFGachaType.STANDARD: "常驻祈愿",
            UIGFGachaType.CHARACTER: "角色活动祈愿",
            UIGFGachaType.WEAPON: "武器活动祈愿",
        }
        data = UIGFListInfo(list=[])
        info = UIGFModel(info=UIGFInfo(export_app=import_type.value), hk4e=[data], hkrpg=[], nap=[])
        if import_type == ImportType.PAIMONMOE:
            ws = wb["Information"]
            if ws["B2"].value != PAIMONMOE_VERSION:
                raise PaimonMoeGachaLogFileError(file_version=ws["B2"].value, support_version=PAIMONMOE_VERSION)
            count = 1
            for gacha_type in paimonmoe_sheets:
                ws = wb[paimonmoe_sheets[gacha_type]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    data.list.append(from_paimon_moe(gacha_type, row[0], row[1], row[2], row[3], count))
                    count += 1
        elif import_type == ImportType.UIGF:
            ws = wb["原始数据"]
            type_map = {}
            count = 0
            for row in ws["1"]:
                if row.value is None:
                    break
                type_map[row.value] = count
                count += 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    break
                data.list.append(
                    from_uigf(
                        row[type_map["uigf_gacha_type"]],
                        row[type_map["gacha_type"]],
                        row[type_map["item_type"]],
                        row[type_map["name"]],
                        row[type_map["time"]],
                        row[type_map["rank_type"]],
                        row[type_map["id"]],
                    )
                )
        else:
            for gacha_type in fxq_sheets:
                ws = wb[fxq_sheets[gacha_type]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    data.list.append(from_fxq(gacha_type, row[2], row[1], row[0], row[3], row[6]))

        return json.loads(info.model_dump_json())
